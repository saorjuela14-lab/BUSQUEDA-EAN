"""
Exportación a Excel con formato corporativo (openpyxl).

Genera un libro .xlsx con varias hojas para los compradores de Makro:
  1. Resumen          — datos del producto y KPIs de mercado.
  2. Precios          — precio regular/promo por retailer y disponibilidad.
  3. Márgenes         — margen $ y % por retailer.
  4. Estrategias      — escenarios de precio Makro.
  5. Alertas          — alertas de negocio detectadas.

Los valores monetarios usan formato de moneda colombiana sin decimales.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import CATEGORIES, REPORTS_DIR

# Paleta corporativa.
_BRAND = "003A8C"        # azul Makro
_ACCENT = "FFC20E"       # amarillo
_HEADER_FILL = PatternFill("solid", fgColor=_BRAND)
_SUBHEADER_FILL = PatternFill("solid", fgColor="E8EDF5")
_WHITE = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(color=_BRAND, bold=True, size=16)
_BOLD = Font(bold=True)
_MONEY_FMT = '"$"#,##0'
_PCT_FMT = "0.0%"
_THIN = Side(style="thin", color="D0D7E2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE
        cell.alignment = _CENTER
        cell.border = _BORDER


def _autofit(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _sheet_summary(wb: Workbook, report: dict) -> None:
    ws = wb.active
    ws.title = "Resumen"
    cat = CATEGORIES.get(report.get("category") or "", {})

    ws.merge_cells("A1:D1")
    ws["A1"] = "Makro Colombia — Inteligencia de Precios"
    ws["A1"].font = _TITLE_FONT

    rows = [
        ("Producto", report.get("product_name")),
        ("EAN", report.get("ean")),
        ("Categoría", cat.get("label", report.get("category"))),
        ("Costo Makro", report.get("cost")),
        ("Modo de búsqueda", "Descripción" if report.get("match_mode") == "description" else "EAN"),
        ("Fecha consulta", report.get("timestamp", datetime.utcnow().isoformat())[:19].replace("T", " ")),
        ("", ""),
        ("KPIs de mercado", ""),
    ]
    kpis = report.get("kpis", {})
    rows += [
        ("Precio mínimo", kpis.get("min_price")),
        ("Precio máximo", kpis.get("max_price")),
        ("Precio promedio", kpis.get("avg_price")),
        ("Spread", kpis.get("spread")),
        ("Retailer líder", kpis.get("leader_retailer")),
        ("Retailer más caro", kpis.get("most_expensive_retailer")),
        ("Retailers con precio", f"{kpis.get('available_count', 0)} de {kpis.get('total_count', 0)}"),
        ("Margen promedio %", kpis.get("avg_margin_pct")),
    ]

    money_labels = {"Costo Makro", "Precio mínimo", "Precio máximo", "Precio promedio", "Spread"}
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = _BOLD
        c = ws.cell(row=r, column=2, value=value)
        if label in money_labels and isinstance(value, (int, float)):
            c.number_format = _MONEY_FMT
        if label in ("KPIs de mercado",):
            ws.cell(row=r, column=1).fill = _SUBHEADER_FILL
        r += 1
    _autofit(ws, {1: 22, 2: 32})


def _sheet_prices(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Precios por retailer")
    headers = ["Retailer", "Disponible", "Precio regular", "Precio promo", "Promoción", "Precio efectivo", "Link"]
    _style_header(ws, 1, headers)
    for i, r in enumerate(report.get("results", []), start=2):
        # Precio efectivo = precio regular (sin descuento); la promo va aparte.
        eff = r.get("price") or r.get("promo_price")
        ws.cell(row=i, column=1, value=r.get("retailer_name") or r.get("retailer"))
        ws.cell(row=i, column=2, value="Sí" if r.get("found") else "No").alignment = _CENTER
        c3 = ws.cell(row=i, column=3, value=r.get("price")); c3.number_format = _MONEY_FMT
        c4 = ws.cell(row=i, column=4, value=r.get("promo_price")); c4.number_format = _MONEY_FMT
        ws.cell(row=i, column=5, value=r.get("promo_desc"))
        c6 = ws.cell(row=i, column=6, value=eff if r.get("found") else None); c6.number_format = _MONEY_FMT
        ws.cell(row=i, column=7, value=r.get("url"))
    _autofit(ws, {1: 16, 2: 11, 3: 15, 4: 14, 5: 20, 6: 15, 7: 40})


def _sheet_margins(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Análisis de márgenes")
    headers = ["Retailer", "Precio regular", "Precio con descuento", "Costo", "Margen $", "Margen %"]
    _style_header(ws, 1, headers)
    cost = report.get("cost")
    for i, m in enumerate(report.get("margins", []), start=2):
        ws.cell(row=i, column=1, value=m.get("retailer"))
        c2 = ws.cell(row=i, column=2, value=m.get("effective_price")); c2.number_format = _MONEY_FMT
        c3 = ws.cell(row=i, column=3, value=m.get("promo_price")); c3.number_format = _MONEY_FMT
        c4 = ws.cell(row=i, column=4, value=cost); c4.number_format = _MONEY_FMT
        c5 = ws.cell(row=i, column=5, value=m.get("margin_value")); c5.number_format = _MONEY_FMT
        pct = m.get("margin_pct")
        c6 = ws.cell(row=i, column=6, value=(pct / 100) if pct is not None else None)
        c6.number_format = _PCT_FMT
    _autofit(ws, {1: 16, 2: 15, 3: 18, 4: 14, 5: 14, 6: 12})


def _sheet_strategies(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Estrategias Makro")
    headers = ["Escenario", "Descripción", "Precio sugerido", "Margen $", "Margen %"]
    _style_header(ws, 1, headers)
    for i, s in enumerate(report.get("strategies", []), start=2):
        ws.cell(row=i, column=1, value=s.get("name")).font = _BOLD
        ws.cell(row=i, column=2, value=s.get("description"))
        c3 = ws.cell(row=i, column=3, value=s.get("suggested_price")); c3.number_format = _MONEY_FMT
        c4 = ws.cell(row=i, column=4, value=s.get("margin_value")); c4.number_format = _MONEY_FMT
        pct = s.get("margin_pct")
        c5 = ws.cell(row=i, column=5, value=(pct / 100) if pct is not None else None)
        c5.number_format = _PCT_FMT
    _autofit(ws, {1: 24, 2: 48, 3: 16, 4: 14, 5: 12})


def _sheet_alerts(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Alertas")
    headers = ["Nivel", "Tipo", "Retailer", "Mensaje"]
    _style_header(ws, 1, headers)
    alerts = report.get("alerts", [])
    if not alerts:
        ws.cell(row=2, column=1, value="Sin alertas").font = _BOLD
    for i, a in enumerate(alerts, start=2):
        ws.cell(row=i, column=1, value=a.get("level"))
        ws.cell(row=i, column=2, value=a.get("type"))
        ws.cell(row=i, column=3, value=a.get("retailer"))
        ws.cell(row=i, column=4, value=a.get("message"))
    _autofit(ws, {1: 10, 2: 16, 3: 14, 4: 70})


def build_workbook(report: dict) -> Workbook:
    """Construye el libro de Excel a partir de un informe de consulta."""
    wb = Workbook()
    _sheet_summary(wb, report)
    _sheet_prices(wb, report)
    _sheet_margins(wb, report)
    _sheet_strategies(wb, report)
    _sheet_alerts(wb, report)
    return wb


def export_report(report: dict, filename: Optional[str] = None) -> Path:
    """Genera y guarda el reporte Excel; devuelve la ruta del archivo."""
    wb = build_workbook(report)
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ean = report.get("ean", "producto")
        filename = f"makro_precios_{ean}_{ts}.xlsx"
    out_path = REPORTS_DIR / filename
    wb.save(out_path)
    return out_path
